"""
Excel/CSV extraction agent.

Flow:
  1. scan_excel()        → identifies sheets and proposes column mappings
  2. extract_from_excel() → extracts records from selected sheets

Key design decisions:
  - Uses openpyxl to resolve merged cells before processing
  - Detects and skips repeated header rows within sheets
  - Extracts document-level metadata (scope, year, country, source) from title/notes rows
  - Processes data in batches of BATCH_SIZE rows to stay within token limits
  - Forward-fills merged cell values so every row has complete data
"""
import asyncio
import json
import pandas as pd
import anthropic
from openpyxl import load_workbook
from app.config import settings
from app.schemas.ingestion import ScanResult, DocumentSection, ExtractedRecord, DocumentMetadata
from app.services.extraction.prompts import EXCEL_SCAN_SYSTEM_PROMPT, METADATA_EXTRACT_PROMPT

client = anthropic.AsyncAnthropic(api_key=settings.anthropic_api_key)

SONNET_INPUT_COST_PER_TOKEN = 3 / 1_000_000
SONNET_OUTPUT_COST_PER_TOKEN = 15 / 1_000_000
BATCH_SIZE = 10      # rows per Claude call
PARALLEL_BATCHES = 1  # sequential — avoids bursting past org rate limits


# ── File reading ──────────────────────────────────────────────────────────────

def _load_sheets_raw(file_path: str) -> dict[str, pd.DataFrame]:
    """
    Load all sheets as raw DataFrames with merged cells fully resolved.
    Returns one DataFrame per sheet where every cell in a merged range
    carries the top-left value (not NaN for the non-first cells).
    """
    if file_path.endswith(".csv"):
        df = pd.read_csv(file_path, header=None, dtype=str)
        return {"Sheet1": df}

    wb = load_workbook(file_path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Build merged-cell value map: (row, col) → value of top-left cell
        merge_map: dict[tuple[int, int], object] = {}
        for mr in ws.merged_cells.ranges:
            val = ws.cell(mr.min_row, mr.min_col).value
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r, c)] = val

        rows = []
        for row in ws.iter_rows():
            rows.append([
                merge_map.get((cell.row, cell.column), cell.value)
                for cell in row
            ])

        if rows:
            result[sheet_name] = pd.DataFrame(rows)

    return result


# ── Sheet preprocessing ───────────────────────────────────────────────────────

EF_HEADER_KEYWORDS = {"activity", "fuel", "unit", "kg", "co2", "emission", "scope", "ghg", "factor"}


def _row_has_numbers(row_values: list) -> bool:
    """Return True if a row contains at least 3 numeric values (data row, not metadata)."""
    count = 0
    for v in row_values:
        if v is None:
            continue
        try:
            float(str(v).replace(",", "").strip())
            count += 1
            if count >= 3:
                return True
        except (ValueError, TypeError):
            pass
    return False


def _find_first_header_row(df: pd.DataFrame) -> int:
    """
    Find the real column-header row by scanning the first 60 rows and looking for
    the LAST short-string row before the first row that contains actual numbers.

    This handles DEFRA-style sheets where:
    - Rows 1–21: title, metadata, guidance text
    - Row 22: real headers ("Activity | Fuel | Unit | kg CO2e | ...")
    - Row 23+: numeric data
    """
    n = min(60, len(df))

    # Find the first row that has real numeric data values
    first_data_row = n  # default: no data found
    for i in range(n):
        if _row_has_numbers(list(df.iloc[i])):
            first_data_row = i
            break

    if first_data_row == 0:
        return 0

    # The header is the last non-empty row before the first data row
    # that has multiple filled columns (i.e. not a title/blank row)
    for i in range(first_data_row - 1, -1, -1):
        row_vals = list(df.iloc[i])
        non_null = [v for v in row_vals if v is not None and str(v).strip()]
        if len(non_null) >= 2:
            return i

    return 0


def _preprocess_sheet(df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """
    Returns (clean_data_df, context_text):
    - clean_data_df: data rows only, with proper column names and forward-filled merged cells
    - context_text: text from all rows before the first header (metadata / guidance)

    Handles two-row headers (e.g. DEFRA: "GHGs" merged above "kg CO2e | kg CO2 | kg CH4...").
    When the row immediately after the first header is also all-strings, the two rows are
    combined: child values take priority, parent values fill gaps.
    """
    header_idx = _find_first_header_row(df)

    # Collect context text from rows before the header
    context_lines = []
    for i in range(header_idx):
        vals = [str(v).strip() for v in df.iloc[i] if v is not None and str(v).strip()]
        if vals:
            context_lines.append(" | ".join(vals))
    context_text = "\n".join(context_lines)

    if header_idx >= len(df) - 1:
        return df, context_text

    # Build column names — handle two-row headers
    raw_headers = list(df.iloc[header_idx])
    data_start_idx = header_idx + 1

    if header_idx + 1 < len(df):
        next_row = list(df.iloc[header_idx + 1])
        next_str_vals = [v for v in next_row if v is not None and str(v).strip()]
        # If the next row is also all non-numeric strings it's a sub-header row
        all_strings = next_str_vals and all(isinstance(v, str) for v in next_str_vals)
        if all_strings:
            # Merge: child value wins; fall back to parent value
            merged = []
            for parent, child in zip(raw_headers, next_row):
                child_s = str(child).strip() if child is not None else ""
                parent_s = str(parent).strip() if parent is not None else ""
                merged.append(child_s if child_s else parent_s)
            raw_headers = merged
            data_start_idx = header_idx + 2
            # Add the parent-header row to context too (it has category info)
            parent_vals = [str(v).strip() for v in list(df.iloc[header_idx]) if v is not None and str(v).strip()]
            if parent_vals:
                context_text = context_text + ("\n" if context_text else "") + " | ".join(parent_vals)

    headers = []
    seen: dict[str, int] = {}
    for h in raw_headers:
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"_col{len(headers)}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    # Slice data rows
    data = df.iloc[data_start_idx:].copy()
    data.columns = headers[: len(data.columns)]
    data = data.reset_index(drop=True)

    # Drop repeated header rows (rows where values match the header names)
    header_lower = {h.lower() for h in headers if not h.startswith("_col")}

    def _is_repeated_header(row) -> bool:
        vals = [str(v).lower().strip() for v in row if v is not None and str(v).strip()]
        return bool(vals) and sum(v in header_lower for v in vals) >= min(3, len(header_lower))

    data = data[~data.apply(_is_repeated_header, axis=1)].copy()

    # Forward-fill merged cells (NaN in Activity/Fuel columns caused by merges)
    for col in data.columns:
        if data[col].notna().any():
            data[col] = data[col].ffill()

    # Drop entirely blank rows (spacer rows between sections)
    data = data.dropna(how="all").reset_index(drop=True)

    # Drop columns that are entirely empty or are spacer columns (_col*)
    keep_cols = [
        c for c in data.columns
        if not c.startswith("_col") and data[c].notna().any()
    ]
    if keep_cols:
        data = data[keep_cols]

    # Drop rows where NO column contains a numeric value — these are FAQ/footnote
    # text rows that appear after the data table (e.g. "I need a conversion factor...")
    def _has_numeric(val) -> bool:
        if isinstance(val, (int, float)):
            return True
        if isinstance(val, str):
            try:
                float(val.replace(",", "").strip())
                return True
            except ValueError:
                pass
        return False

    has_number = data.apply(lambda row: any(_has_numeric(v) for v in row), axis=1)
    data = data[has_number].reset_index(drop=True)

    return data, context_text


# ── JSON safety ───────────────────────────────────────────────────────────────

def _df_to_records(df: pd.DataFrame) -> list:
    """Convert DataFrame to a guaranteed JSON-safe list of dicts."""
    df = df.copy()
    for col in df.columns:
        if hasattr(df[col], "dt") or str(df[col].dtype).startswith("datetime"):
            df[col] = df[col].astype(str)
    raw = df.fillna("").to_dict(orient="records")
    return json.loads(json.dumps(raw, default=str))


# ── Scan ──────────────────────────────────────────────────────────────────────

def _sheet_is_cover_or_notes(raw_df: pd.DataFrame) -> bool:
    """Return True if a sheet looks like a cover page / notes sheet (no meaningful numeric data)."""
    numeric_count = 0
    for _, row in raw_df.iterrows():
        for v in row:
            if v is None:
                continue
            try:
                float(str(v).replace(",", "").strip())
                numeric_count += 1
                if numeric_count >= 5:
                    return False
            except (ValueError, TypeError):
                pass
    return True  # fewer than 5 numeric values → treat as cover/notes


def _extract_sheet_as_text(raw_df: pd.DataFrame, max_chars: int = 5000) -> str:
    """Extract all readable text from a sheet (for cover pages / notes)."""
    lines = []
    for _, row in raw_df.iterrows():
        cells = [
            str(v).strip()
            for v in row
            if v is not None and str(v).strip() not in ("", "None", "nan")
        ]
        if cells:
            lines.append("  |  ".join(cells))
    return "\n".join(lines)[:max_chars]


async def _extract_document_metadata(
    context_text: str,
    cover_text: str,
    preview_rows: list,
) -> DocumentMetadata:
    """Call Claude to extract document-level metadata from all available text sources."""
    if not context_text and not cover_text and not preview_rows:
        return DocumentMetadata()

    content = ""
    if cover_text:
        content += (
            "COVER PAGE / NOTES SHEETS — read carefully for validity dates, "
            "LCA stages, applicability guidance, and scope information:\n"
            f"{cover_text}\n\n"
        )
    if context_text:
        content += f"DATA SHEET HEADERS / CONTEXT:\n{context_text}\n\n"
    if preview_rows:
        content += f"SAMPLE DATA ROWS:\n{json.dumps(preview_rows[:10], indent=2, default=str)}"

    try:
        response = await client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            system=METADATA_EXTRACT_PROMPT,
            messages=[{"role": "user", "content": content}],
        )
        raw = response.content[0].text
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start >= 0 and end > start:
            data = json.loads(raw[start:end])
            return DocumentMetadata(**{k: v for k, v in data.items() if v is not None})
    except Exception:
        pass
    return DocumentMetadata()


async def scan_excel(file_path: str, document_id: str, session_id: str) -> ScanResult:
    """Scan an Excel/CSV file: resolve merges, extract context, identify EF sheets."""
    raw_sheets = _load_sheets_raw(file_path)

    sections = []
    total_rows = 0
    all_context_text = ""
    all_cover_text = ""      # full text from cover/notes sheets
    all_preview_rows: list = []

    for i, (sheet_name, raw_df) in enumerate(raw_sheets.items()):
        # Cover/notes sheets: extract full text for metadata, skip as data section
        if _sheet_is_cover_or_notes(raw_df):
            sheet_text = _extract_sheet_as_text(raw_df)
            if sheet_text:
                all_cover_text += f"\n[Sheet: {sheet_name}]\n{sheet_text}\n"
            sections.append(DocumentSection(
                index=i,
                title=sheet_name,
                page_range="cover/notes",
                column_headers=[],
                description="Cover page or notes sheet — metadata extracted automatically, no EF rows.",
                row_count_estimate=0,
            ))
            continue

        data_df, context_text = _preprocess_sheet(raw_df)
        total_rows += len(data_df)
        if context_text:
            all_context_text += f"\n[Sheet: {sheet_name}]\n{context_text}"

        has_numeric = any(
            pd.to_numeric(data_df[c], errors="coerce").notna().any()
            for c in data_df.columns
        )
        if not has_numeric and len(data_df) == 0:
            continue

        preview = _df_to_records(data_df.head(5))
        all_preview_rows.extend(preview)

        description = (
            f"Columns: {list(data_df.columns)[:10]}. "
            f"{len(data_df)} data rows after preprocessing."
        )
        if context_text:
            description += f" Context: {context_text[:200]}"

        sections.append(DocumentSection(
            index=i,
            title=sheet_name,
            page_range=f"rows 1–{len(data_df)}",
            column_headers=[str(c) for c in data_df.columns],
            description=description,
            row_count_estimate=len(data_df),
        ))

    # Extract document-level metadata — cover pages provide the richest context
    document_metadata = await _extract_document_metadata(
        all_context_text, all_cover_text, all_preview_rows
    )

    estimated_tokens = total_rows * 50 + len(all_context_text) // 4
    estimated_cost = (
        estimated_tokens * SONNET_INPUT_COST_PER_TOKEN +
        500 * SONNET_OUTPUT_COST_PER_TOKEN
    )

    return ScanResult(
        session_id=session_id,
        document_id=document_id,
        sections_found=sections,
        estimated_tokens=estimated_tokens,
        estimated_cost_usd=round(estimated_cost, 4),
        page_count=total_rows,
        has_scanned_pages=False,
        document_metadata=document_metadata,
    )


# ── Extract ───────────────────────────────────────────────────────────────────

EXCEL_EXTRACT_SYSTEM = """You are a GHG emission factor extraction specialist working with spreadsheet data.

You receive JSON rows from a cleaned spreadsheet. Each row is a dict of {column_name: cell_value}.
Your task: produce ONE output record per input row.

COLUMN → SCHEMA FIELD MAPPING RULES:
Map spreadsheet column names to schema fields using these patterns:
  ef_total_co2e  ← columns whose name contains "CO2e" or "CO2eq" or "GHG" (the total combined factor)
  ef_co2         ← columns whose name contains "CO2" AND also contains "CO2" component specifically (not the total)
  ef_ch4         ← columns whose name contains "CH4" or "methane"
  ef_n2o         ← columns whose name contains "N2O" or "nitrous"
  ef_pfc         ← columns whose name contains "PFC" or "HFC"
  ef_sf6         ← columns whose name contains "SF6"
  source_activity_name ← columns named "Activity", "Fuel", "Fuel type", "Item", "Description", "Source"
  unit           ← columns named "Unit", "Units"

IMPORTANT: When a column called "kg CO2e" exists and has a number, that number IS the ef_total_co2e value.
When "kg CO2e of CO2 per unit" exists, that IS the ef_co2 value. And so on.
NEVER return null for a numeric field when the input row contains a non-empty number for it.

APPLY DOCUMENT CONTEXT to every record: geography_country, gwp_version, applicable_scopes, source_name, source_type, validity_start, validity_end from the CONFIRMED METADATA block.

OUTPUT FORMAT — return a JSON array. Each element is one record with this EXACT structure:
[
  {
    "source_activity_name":   {"value": "Butane", "source_snippet": "Butane", "extraction_confidence": "high"},
    "canonical_activity_name":{"value": "Butane combustion", "source_snippet": "Butane", "extraction_confidence": "high"},
    "activity_category":      {"value": "Fuel combustion > Gaseous fuels > Butane", "source_snippet": null, "extraction_confidence": "high"},
    "unit":                   {"value": "tonnes", "source_snippet": "tonnes", "extraction_confidence": "high"},
    "ef_total_co2e":          {"value": 3033.38, "source_snippet": "3033.3806711409397", "extraction_confidence": "high"},
    "ef_co2":                 {"value": 3029.26, "source_snippet": "3029.26", "extraction_confidence": "high"},
    "ef_ch4":                 {"value": 2.52, "source_snippet": "2.52", "extraction_confidence": "high"},
    "ef_n2o":                 {"value": 1.60, "source_snippet": "1.6006711409395973", "extraction_confidence": "high"},
    "ef_pfc":                 {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "ef_sf6":                 {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "ef_nf3":                 {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "applicable_scopes":      {"value": ["Scope 1"], "source_snippet": null, "extraction_confidence": "high"},
    "lca_stages":             {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "source_name":            {"value": "UK Government GHG Conversion Factors 2023", "source_snippet": null, "extraction_confidence": "high"},
    "source_type":            {"value": "Government / Regulatory body", "source_snippet": null, "extraction_confidence": "high"},
    "source_url":             {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "validity_start":         {"value": 2023, "source_snippet": null, "extraction_confidence": "high"},
    "validity_end":           {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "geography_global":       {"value": false, "source_snippet": null, "extraction_confidence": "high"},
    "geography_country":      {"value": "GB", "source_snippet": null, "extraction_confidence": "high"},
    "geography_region":       {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "gwp_version":            {"value": "AR5", "source_snippet": null, "extraction_confidence": "high"},
    "supplier_name":          {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "supplier_country":       {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "supplier_sector":        {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "supplier_epd_reference": {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "comments_applicability": {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "comments_limitations":   {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "custom_tags":            {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "additional_notes":       {"value": null, "source_snippet": null, "extraction_confidence": "high"},
    "has_outlier_values": false,
    "has_unit_mismatch": false,
    "outlier_notes": []
  }
]

Return ONLY the JSON array. No markdown fences. No explanatory text. No truncation."""


async def extract_from_excel(file_path: str, section_indices: list[int], confirmed_metadata: DocumentMetadata | None = None) -> list[ExtractedRecord]:
    """
    Extract EF records from selected sheets.
    - Resolves merged cells via openpyxl
    - Strips non-data rows and repeated headers
    - Extracts document context and passes it to Claude with every batch
    - Processes in BATCH_SIZE-row batches to avoid token truncation
    """
    from app.services.extraction.pdf_agent import _parse_extraction_response

    raw_sheets = _load_sheets_raw(file_path)
    sheet_names = list(raw_sheets.keys())

    all_records = []
    record_index = 0

    for section_idx in section_indices:
        if section_idx >= len(sheet_names):
            continue

        sheet_name = sheet_names[section_idx]
        raw_df = raw_sheets[sheet_name]
        data_df, context_text = _preprocess_sheet(raw_df)
        rows_json = _df_to_records(data_df)

        # Build context block: confirmed metadata takes priority over auto-detected context
        meta_lines = []
        if confirmed_metadata:
            m = confirmed_metadata
            if m.source_name:         meta_lines.append(f"Source name: {m.source_name}")
            if m.source_type:         meta_lines.append(f"Source type: {m.source_type}")
            if m.geography_country:   meta_lines.append(f"Geography: {m.geography_description or ''} ({m.geography_country})")
            if m.gwp_version:         meta_lines.append(f"GWP version: {m.gwp_version}")
            if m.applicable_scopes:   meta_lines.append(f"Applicable scopes: {', '.join(m.applicable_scopes)}")
            if m.lca_stages:          meta_lines.append(f"LCA stages: {', '.join(m.lca_stages)}")
            # Validity period (prefer explicit start/end years over single publication year)
            if m.validity_start or m.validity_end:
                start_str = str(m.validity_start) if m.validity_start else "unknown"
                end_str = str(m.validity_end) if m.validity_end else "present"
                meta_lines.append(f"Validity period: {start_str}–{end_str} (use {start_str}-01-01 as validity_start and {end_str + '-12-31' if m.validity_end else 'null'} as validity_end)")
            elif m.year:
                meta_lines.append(f"Publication year: {m.year} (use {m.year}-01-01 as validity_start)")
            if m.comments_applicability: meta_lines.append(f"Applicability (add to every record's comments_applicability): {m.comments_applicability}")
            if m.guidance_notes:         meta_lines.append(f"Additional guidance: {m.guidance_notes}")

        confirmed_block = (
            "CONFIRMED DOCUMENT METADATA (apply ALL of these to every record):\n"
            + "\n".join(meta_lines) + "\n\n"
        ) if meta_lines else ""

        raw_context_block = (
            f"ADDITIONAL DOCUMENT CONTEXT:\n{context_text}\n\n"
            if context_text else ""
        )
        context_block = confirmed_block + raw_context_block

        n_batches = -(-len(rows_json) // BATCH_SIZE)
        print(f"[extract] Sheet '{sheet_name}': {len(rows_json)} rows → {n_batches} batches", flush=True)

        for batch_num in range(n_batches):
            batch_start = batch_num * BATCH_SIZE
            batch = rows_json[batch_start: batch_start + BATCH_SIZE]
            print(f"[extract]   [{sheet_name}] batch {batch_num+1}/{n_batches}: rows {batch_start+1}–{batch_start+len(batch)}", flush=True)
            msg = (
                f"{context_block}"
                f"DATA — Sheet: {sheet_name} | "
                f"Rows {batch_start + 1}–{batch_start + len(batch)} of {len(rows_json)}\n"
                f"Columns: {[str(c) for c in data_df.columns]}\n\n"
                f"{json.dumps(batch, indent=2, default=str)}\n\n"
                "Extract all emission factor records from this batch. "
                "Apply the document context to every record. "
                "Return a JSON array — one element per data row."
            )
            # Retry up to 4 times on rate-limit (429) errors with exponential backoff
            for attempt in range(4):
                try:
                    resp = await client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=16000,
                        system=EXCEL_EXTRACT_SYSTEM,
                        messages=[{"role": "user", "content": msg}],
                    )
                    batch_records = _parse_extraction_response(resp.content[0].text)
                    for rec in batch_records:
                        rec.index = record_index
                        record_index += 1
                    all_records.extend(batch_records)
                    # Pause between batches to stay within per-minute token rate limits
                    await asyncio.sleep(15)
                    break
                except Exception as e:
                    if "429" in str(e) and attempt < 3:
                        wait = 60 * (2 ** attempt)  # 60s, 120s, 240s
                        print(f"[extract]   Rate limited, waiting {wait}s (attempt {attempt+1})", flush=True)
                        await asyncio.sleep(wait)
                    else:
                        raise

    return all_records
